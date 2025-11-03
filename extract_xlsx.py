from openpyxl import load_workbook
import sys
import json

def extract_xlsx(file_path):
    wb = load_workbook(file_path, data_only=True)
    content = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_data = []

        for row in sheet.iter_rows(values_only=True):
            # Only include rows that have at least one non-None value
            if any(cell is not None for cell in row):
                sheet_data.append([str(cell) if cell is not None else '' for cell in row])

        content[sheet_name] = sheet_data

    return content

if __name__ == '__main__':
    file_path = sys.argv[1]
    content = extract_xlsx(file_path)
    print(json.dumps(content, ensure_ascii=False, indent=2))
